from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.db import transaction, models
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from api.models.archive_record_models import ArchiveRecord
from api.models.user_models import User
from api.models.thesis_models import Thesis
from api.models.document_models import Document
from api.models.group_models import Group
from api.serializers.archive_serializers import ArchiveRecordSerializer
from api.permissions.role_permissions import IsAdmin, IsAdviser

class ArchiveRecordViewSet(viewsets.ModelViewSet):
    queryset = ArchiveRecord.objects.all().select_related('archived_by')
    serializer_class = ArchiveRecordSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Admins can see all archive records
        if self.request.user.role == 'ADMIN':
            return queryset
            
        # Advisers can see archive records for their advised theses
        elif self.request.user.role == 'ADVISER':
            return queryset.filter(
                content_type='thesis',
                archived_by__advised_theses__adviser=self.request.user
            ) | queryset.filter(archived_by=self.request.user)
            
        # Students can only see archive records they created
        elif self.request.user.role == 'STUDENT':
            return queryset.filter(archived_by=self.request.user)
            
        # Panels can only see archive records they created
        elif self.request.user.role == 'PANEL':
            return queryset.filter(archived_by=self.request.user)
            
        return queryset.none()

    def perform_create(self, serializer):
        serializer.save(archived_by=self.request.user)

    @action(detail=False, methods=['get'])
    def thesis_archives(self, request):
        """Get all thesis archive records"""
        thesis_archives = ArchiveRecord.objects.filter(content_type='thesis').select_related('archived_by')
        serializer = self.get_serializer(thesis_archives, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def document_archives(self, request):
        """Get all document archive records"""
        document_archives = ArchiveRecord.objects.filter(content_type='document').select_related('archived_by')
        serializer = self.get_serializer(document_archives, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def group_archives(self, request):
        """Get all group archive records"""
        group_archives = ArchiveRecord.objects.filter(content_type='group').select_related('archived_by')
        serializer = self.get_serializer(group_archives, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def restore(self, request, pk=None):
        """Restore an archived record (admin only)"""
        archive_record = self.get_object()
        
        # For now, we'll just mark it as restored in the archive record
        # In a real implementation, you would restore the actual data
        return Response(
            {'detail': 'Archive record marked for restoration. Actual restoration would be implemented in a real system.'},
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=['post'])
    def archive_thesis(self, request):
        """Archive a thesis"""
        thesis_id = request.data.get('thesis_id')
        reason = request.data.get('reason', '')
        retention_period = request.data.get('retention_period_years', 7)
        
        if not thesis_id:
            return Response(
                {'detail': 'thesis_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            thesis = Thesis.objects.get(id=thesis_id)
        except Thesis.DoesNotExist:
            return Response(
                {'detail': 'Invalid thesis ID'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # Check permissions - only admins and advisers can archive theses
        if not (self.request.user.role == 'ADMIN' or 
                (self.request.user.role == 'ADVISER' and thesis.adviser == self.request.user)):
            return Response(
                {'detail': 'You do not have permission to archive this thesis'},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # Get panel members
        panel_members = []
        if thesis.group and thesis.group.panels:
            panel_members = [f"{panel.first_name} {panel.last_name}" for panel in thesis.group.panels.all()]

        # Get adviser name
        adviser_name = None
        if thesis.adviser:
            adviser_name = f"{thesis.adviser.first_name} {thesis.adviser.last_name}"

        # Get keywords as list
        keywords_list = thesis.get_keywords_list()

        # Create archive record with all required information
        archive_data = {
            'title': thesis.title,
            'abstract': thesis.abstract,
            'keywords': keywords_list,
            'status': thesis.status,
            'adviser': str(thesis.adviser.id) if thesis.adviser else None,
            'adviser_name': adviser_name,
            'group': str(thesis.group.id),
            'group_name': thesis.group.name if thesis.group else 'Unknown Group',
            'panels': panel_members,
            'drive_folder_url': thesis.get_drive_folder_url(),
            'finished_at': timezone.now().isoformat(),
            'created_at': thesis.created_at.isoformat(),
            'updated_at': thesis.updated_at.isoformat(),
        }
        
        archive_record = ArchiveRecord.objects.create(
            content_type='thesis',
            original_id=thesis.id,
            data=archive_data,
            archived_by=self.request.user,
            reason=reason,
            retention_period_years=retention_period
        )
        
        # Update Google Drive folder permissions to read-only if folder exists
        if thesis.drive_folder_id:
            try:
                from api.services.google_drive_service import GoogleDriveService
                # Initialize Google Drive service with the thesis proposer's credentials if available
                drive_service = GoogleDriveService(user=thesis.proposer if thesis.proposer else None)
                if drive_service.service:
                    success = drive_service.update_folder_permissions_to_readonly(thesis.drive_folder_id)
                    if success:
                        print(f"Successfully updated Google Drive folder {thesis.drive_folder_id} to read-only")
                    else:
                        print(f"Failed to update Google Drive folder {thesis.drive_folder_id} to read-only")
                else:
                    print(f"Google Drive service not available for folder {thesis.drive_folder_id}")
            except Exception as e:
                print(f"Error updating Google Drive folder permissions: {e}")
        
        serializer = self.get_serializer(archive_record)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def archive_document(self, request):
        """Archive a document"""
        document_id = request.data.get('document_id')
        reason = request.data.get('reason', '')
        retention_period = request.data.get('retention_period_years', 7)
        
        if not document_id:
            return Response(
                {'detail': 'document_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            document = Document.objects.get(id=document_id)
        except Document.DoesNotExist:
            return Response(
                {'detail': 'Invalid document ID'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # Check permissions - only admins, advisers, and owners can archive documents
        if not (self.request.user.role == 'ADMIN' or 
                (self.request.user.role == 'ADVISER' and document.thesis.adviser == self.request.user) or
                document.uploaded_by == self.request.user):
            return Response(
                {'detail': 'You do not have permission to archive this document'},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # Create archive record
        archive_data = {
            'title': document.get_document_type_display(),
            'document_type': document.document_type,
            'file_path': str(document.file) if document.file else None,
            'google_doc_id': document.google_doc_id,
            'provider': document.provider,
            'uploaded_by': str(document.uploaded_by.id) if document.uploaded_by else None,
            'thesis': str(document.thesis.id) if document.thesis else None,
            'created_at': document.created_at.isoformat(),
            'updated_at': document.updated_at.isoformat(),
        }
        
        archive_record = ArchiveRecord.objects.create(
            content_type='document',
            original_id=document.id,
            data=archive_data,
            archived_by=self.request.user,
            reason=reason,
            retention_period_years=retention_period
        )
        
        serializer = self.get_serializer(archive_record)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def archive_group(self, request):
        """Archive a group"""
        group_id = request.data.get('group_id')
        reason = request.data.get('reason', '')
        retention_period = request.data.get('retention_period_years', 7)
        
        if not group_id:
            return Response(
                {'detail': 'group_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            group = Group.objects.get(id=group_id)
        except Group.DoesNotExist:
            return Response(
                {'detail': 'Invalid group ID'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # Check permissions - only admins and advisers can archive groups
        if not (self.request.user.role == 'ADMIN' or
                (self.request.user.role == 'ADVISER' and group.adviser == self.request.user)):
            return Response(
                {'detail': 'You do not have permission to archive this group'},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # Create archive record
        archive_data = {
            'name': group.name,
            'abstract': group.abstract,
            'keywords': group.keywords,
            'possible_topics': group.possible_topics,
            'status': group.status,
            'adviser': str(group.adviser.id) if group.adviser else None,
            'leader': str(group.leader.id) if group.leader else None,
            'members': [str(member.id) for member in group.members.all()],
            'panels': [str(panel.id) for panel in group.panels.all()],
            'created_at': group.created_at.isoformat(),
            'updated_at': group.updated_at.isoformat(),
        }
        
        archive_record = ArchiveRecord.objects.create(
            content_type='group',
            original_id=group.id,
            data=archive_data,
            archived_by=self.request.user,
            reason=reason,
            retention_period_years=retention_period
        )
        
        serializer = self.get_serializer(archive_record)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def search(self, request):
        """Search archive records by keywords and filter by year"""
        query = request.GET.get('q', '').strip()
        year = request.GET.get('year', '').strip()
        
        # Start with the base queryset based on user permissions
        queryset = self.get_queryset().filter(content_type='thesis')
        
        # Apply search query if provided
        if query:
            queryset = queryset.filter(
                Q(data__title__icontains=query) |
                Q(data__abstract__icontains=query) |
                Q(data__keywords__icontains=query)
            )
        
        # Apply year filter if provided
        if year and year != 'all':
            try:
                year_int = int(year)
                start_date = timezone.make_aware(timezone.datetime(year_int, 1, 1))
                end_date = timezone.make_aware(timezone.datetime(year_int + 1, 1, 1))
                queryset = queryset.filter(
                    archived_at__gte=start_date,
                    archived_at__lt=end_date
                )
            except (ValueError, TypeError):
                pass  # Invalid year, ignore filter
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def download_pdf_report(self, request, year=None):
        """Download PDF report of finished theses for a specific year"""
        if year is None:
            year = request.GET.get('year')
        if not year:
            return Response(
                {'detail': 'Year parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            year_int = int(year)
        except ValueError:
            return Response(
                {'detail': 'Invalid year format'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check permissions - only admins and advisers can download reports
        if not hasattr(request, 'user') or not request.user.is_authenticated:
            return Response(
                {'detail': 'Authentication required'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        if request.user.role not in ['ADMIN', 'ADVISER']:
            return Response(
                {'detail': 'You do not have permission to download reports'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get archived theses for the specified year
        start_date = timezone.make_aware(timezone.datetime(year_int, 1, 1))
        end_date = timezone.make_aware(timezone.datetime(year_int + 1, 1, 1))

        print(f"DEBUG: Filtering archives for year {year_int}")
        print(f"DEBUG: Start date: {start_date}")
        print(f"DEBUG: End date: {end_date}")
        print(f"DEBUG: User role: {request.user.role}")

        # Filter based on user role
        queryset = ArchiveRecord.objects.filter(
            content_type='thesis',
            archived_at__gte=start_date,
            archived_at__lt=end_date
        ).filter(
            models.Q(data__status='FINAL_APPROVED') | models.Q(data__status='ARCHIVED')
        )  # Include both finalized approved theses and archived theses

        print(f"DEBUG: Base queryset count: {queryset.count()}")

        # Apply adviser filtering if needed
        if request.user.role == 'ADVISER':
            # Advisers can only see theses they advised
            queryset = queryset.filter(
                data__adviser=str(request.user.id)
            )
            print(f"DEBUG: Adviser filtered queryset count: {queryset.count()}")

        print(f"DEBUG: Final queryset count: {queryset.count()}")

        # Create PDF document
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )

        normal_style = styles['Normal']

        story = []

        # Title
        title_text = f"Thesis Archive Report - {year}"
        title = Paragraph(title_text, title_style)
        story.append(title)
        story.append(Spacer(1, 12))

        # Report info
        info_text = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}\nTotal theses: {queryset.count()}"
        info = Paragraph(info_text, normal_style)
        story.append(info)
        story.append(Spacer(1, 20))

        if queryset.exists():
            # Simple text-based report instead of table
            for archive in queryset:
                archive_data = archive.data or {}

                # Extract data with safe defaults
                group_name = str(archive_data.get('group_name', 'Unknown Group'))
                group_members = archive_data.get('group_members', [])
                group_members_str = ', '.join(str(m) for m in group_members) if group_members else 'No members listed'
                topic = str(archive_data.get('title', 'Unknown Topic'))
                abstract = str(archive_data.get('abstract', 'No abstract available'))
                finished_at = archive.archived_at.strftime('%Y-%m-%d %H:%M:%S') if archive.archived_at else 'Unknown'
                panels = archive_data.get('panels', [])
                panel_names = ', '.join(str(p) for p in panels) if panels else 'No panel assigned'

                # Add thesis info
                thesis_info = f"""Group: {group_name}
Group Members: {group_members_str}
Topic: {topic}
Abstract: {abstract[:300]}{'...' if len(abstract) > 300 else ''}
Finished: {finished_at}
Panel: {panel_names}
"""
                story.append(Paragraph(thesis_info, normal_style))
                story.append(Spacer(1, 12))
        else:
            # No data message
            no_data = Paragraph("No archived theses found for the selected year.", normal_style)
            story.append(no_data)

        # Ensure we have at least one element in the story
        if not story:
            story.append(Paragraph("Error: Unable to generate report.", normal_style))

        # Build PDF
        doc.build(story)
        buffer.seek(0)
        pdf_data = buffer.getvalue()
        buffer.close()

        response = HttpResponse(
            pdf_data,
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename=thesis_report_{year}.pdf'
        response['Content-Length'] = len(pdf_data)

        return response

    def download_excel_report(self, request, year=None):
        """Download Excel report of finished theses for a specific year"""
        if year is None:
            year = request.GET.get('year')
        if not year:
            return Response(
                {'detail': 'Year parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            year_int = int(year)
        except ValueError:
            return Response(
                {'detail': 'Invalid year format'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check permissions - only admins can download Excel reports
        if not hasattr(request, 'user') or not request.user.is_authenticated:
            return Response(
                {'detail': 'Authentication required'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        if request.user.role != 'ADMIN':
            return Response(
                {'detail': 'Only admin users can download Excel reports'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get archived theses for the specified year
        start_date = timezone.make_aware(timezone.datetime(year_int, 1, 1))
        end_date = timezone.make_aware(timezone.datetime(year_int + 1, 1, 1))

        # Filter based on user role
        queryset = ArchiveRecord.objects.filter(
            content_type='thesis',
            archived_at__gte=start_date,
            archived_at__lt=end_date
        ).filter(
            models.Q(data__status='FINAL_APPROVED') | models.Q(data__status='ARCHIVED')
        )  # Include both finalized approved theses and archived theses        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(f"Thesis Report {year}")
        else:
            ws.title = f"Thesis Report {year}"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Headers - Updated to match requirements
        headers = ['Thesis Title', 'Abstract', 'Keywords', 'Group Name', 'Adviser Name', 'Panel Member Names']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Data rows
        row_num = 2
        if queryset.exists():
            for archive in queryset:
                data = archive.data or {}

                # Extract thesis title
                title = str(data.get('title', 'Unknown Title'))

                # Extract abstract
                abstract = str(data.get('abstract', 'No abstract available'))

                # Extract keywords
                keywords_list = data.get('keywords', [])
                keywords = ', '.join(str(k) for k in keywords_list) if keywords_list else 'No keywords'

                # Extract group name
                group_name = str(data.get('group_name', 'Unknown Group'))

                # Extract adviser name
                adviser_name = str(data.get('adviser_name', 'No adviser assigned'))

                # Extract panel members
                panels = data.get('panels', [])
                panel_names = ', '.join(str(p) for p in panels) if panels else 'No panel assigned'

                # Write data to Excel
                ws.cell(row=row_num, column=1, value=title)
                ws.cell(row=row_num, column=2, value=abstract)
                ws.cell(row=row_num, column=3, value=keywords)
                ws.cell(row=row_num, column=4, value=group_name)
                ws.cell(row=row_num, column=5, value=adviser_name)
                ws.cell(row=row_num, column=6, value=panel_names)

                row_num += 1
        else:
            # Add a row indicating no data
            ws.cell(row=row_num, column=1, value='No archived theses found for the selected year')
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create response with BytesIO buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_data = buffer.getvalue()
        buffer.close()

        response = HttpResponse(
            file_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=thesis_report_{year}.xlsx'
        response['Content-Length'] = len(file_data)

        return response

    def download_doc_report(self, request, year=None):
        """Download DOC (HTML) report of finished theses for a specific year"""
        if year is None:
            year = request.GET.get('year')
        if not year:
            return Response(
                {'detail': 'Year parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            year_int = int(year)
        except ValueError:
            return Response(
                {'detail': 'Invalid year format'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check permissions - only admins and advisers can download reports
        if not hasattr(request, 'user') or not request.user.is_authenticated:
            return Response(
                {'detail': 'Authentication required'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        if request.user.role not in ['ADMIN', 'ADVISER']:
            return Response(
                {'detail': 'You do not have permission to download reports'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get archived theses for the specified year
        start_date = timezone.make_aware(timezone.datetime(year_int, 1, 1))
        end_date = timezone.make_aware(timezone.datetime(year_int + 1, 1, 1))

        print(f"DEBUG: Filtering archives for year {year_int}")
        print(f"DEBUG: Start date: {start_date}")
        print(f"DEBUG: End date: {end_date}")
        print(f"DEBUG: User role: {request.user.role}")

        # Filter based on user role
        queryset = ArchiveRecord.objects.filter(
            content_type='thesis',
            archived_at__gte=start_date,
            archived_at__lt=end_date
        ).filter(
            models.Q(data__status='FINAL_APPROVED') | models.Q(data__status='ARCHIVED')
        )  # Include both finalized approved theses and archived theses

        print(f"DEBUG: Base queryset count: {queryset.count()}")
        # Apply adviser filtering if needed
        if request.user.role == 'ADVISER':
            # Advisers can only see theses they advised
            queryset = queryset.filter(
                data__adviser=str(request.user.id)
            )
            print(f"DEBUG: Adviser filtered queryset count: {queryset.count()}")

        print(f"DEBUG: Final queryset count: {queryset.count()}")

        # Create HTML content that can be opened as DOC
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Thesis Archive Report - {year}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #333; text-align: center; }}
        .thesis {{ margin-bottom: 30px; border-bottom: 1px solid #ccc; padding-bottom: 20px; }}
        .field {{ margin-bottom: 10px; }}
        .label {{ font-weight: bold; color: #666; }}
        .value {{ margin-left: 10px; }}
        .metadata {{ text-align: center; color: #666; margin-bottom: 30px; }}
    </style>
</head>
<body>
    <h1>Thesis Archive Report - {year}</h1>
    <div class="metadata">
        Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
        Total theses: {queryset.count()}
    </div>
"""

        if queryset.exists():
            for archive in queryset:
                data = archive.data or {}

                group_name = str(data.get('group_name', 'Unknown Group'))
                group_members = data.get('group_members', [])
                group_members_str = ', '.join(str(m) for m in group_members) if group_members else 'No members listed'
                topic = str(data.get('title', 'Unknown Topic'))
                abstract = str(data.get('abstract', 'No abstract available'))
                finished_at = archive.archived_at.strftime('%Y-%m-%d %H:%M:%S') if archive.archived_at else 'Unknown'
                panels = data.get('panels', [])
                panel_names = ', '.join(str(p) for p in panels) if panels else 'No panel assigned'

                html_content += f"""
    <div class="thesis">
        <div class="field">
            <span class="label">Group:</span>
            <span class="value">{group_name}</span>
        </div>
        <div class="field">
            <span class="label">Group Members:</span>
            <span class="value">{group_members_str}</span>
        </div>
        <div class="field">
            <span class="label">Topic:</span>
            <span class="value">{topic}</span>
        </div>
        <div class="field">
            <span class="label">Abstract:</span>
            <span class="value">{abstract}</span>
        </div>
        <div class="field">
            <span class="label">Finished:</span>
            <span class="value">{finished_at}</span>
        </div>
        <div class="field">
            <span class="label">Panel:</span>
            <span class="value">{panel_names}</span>
        </div>
    </div>
"""
        else:
            html_content += """
    <div class="thesis">
        <p>No archived theses found for the selected year.</p>
    </div>
"""

        html_content += """
</body>
</html>
"""

        response = HttpResponse(
            html_content,
            content_type='application/msword'
        )
        response['Content-Disposition'] = f'attachment; filename=thesis_report_{year}.doc'
        response['Content-Length'] = len(html_content.encode('utf-8'))

        return response

    @action(detail=False, methods=['post'], url_path='download_report', permission_classes=[permissions.IsAuthenticated])
    def download_report(self, request):
        """Download report of finished theses for a specific year in selected format"""
        year = request.data.get('year')
        format_type = request.data.get('format', 'pdf')  # Default to PDF


        if not year:
            return Response(
                {'detail': 'Year is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if format_type not in ['pdf', 'excel', 'doc']:
            return Response(
                {'detail': 'Invalid format. Choose pdf, excel, or doc'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check permissions - only admins can download Excel reports
        if format_type == 'excel' and request.user.role != 'ADMIN':
            return Response(
                {'detail': 'Only admin users can download Excel reports'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Call the appropriate method directly with the year parameter
        if format_type == 'pdf':
            return self.download_pdf_report(request, year)
        elif format_type == 'excel':
            return self.download_excel_report(request, year)
        else:  # doc
            return self.download_doc_report(request, year)

    @action(detail=False, methods=['post'])
    def archive_document(self, request):
        """Archive a document"""
        document_id = request.data.get('document_id')
        reason = request.data.get('reason', '')
        retention_period = request.data.get('retention_period_years', 7)
        
        if not document_id:
            return Response(
                {'detail': 'document_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            document = Document.objects.get(id=document_id)
        except Document.DoesNotExist:
            return Response(
                {'detail': 'Invalid document ID'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # Check permissions - only admins, advisers, and owners can archive documents
        if not (self.request.user.role == 'ADMIN' or 
                (self.request.user.role == 'ADVISER' and document.thesis.adviser == self.request.user) or
                document.uploaded_by == self.request.user):
            return Response(
                {'detail': 'You do not have permission to archive this document'},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # Create archive record
        archive_data = {
            'title': document.get_document_type_display(),
            'document_type': document.document_type,
            'file_path': str(document.file) if document.file else None,
            'google_doc_id': document.google_doc_id,
            'provider': document.provider,
            'uploaded_by': str(document.uploaded_by.id) if document.uploaded_by else None,
            'thesis': str(document.thesis.id) if document.thesis else None,
            'created_at': document.created_at.isoformat(),
            'updated_at': document.updated_at.isoformat(),
        }
        
        archive_record = ArchiveRecord.objects.create(
            content_type='document',
            original_id=document.id,
            data=archive_data,
            archived_by=self.request.user,
            reason=reason,
            retention_period_years=retention_period
        )
        
        serializer = self.get_serializer(archive_record)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
